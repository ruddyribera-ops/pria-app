"""
XlsxExportService - Generates XLSX (Excel) documents from PDC and weekly plans in PRIA v7
"""
from typing import List, Optional
from datetime import datetime
from app.models.pdc import PDC
from app.schemas.export import BrandingConfig

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


def hex_to_rgb(hex_color: str) -> tuple:
    """Convert hex color (#RRGGBB) to RGB hex string for openpyxl"""
    hex_color = hex_color.lstrip("#")
    return hex_color.upper()


class XlsxExportService:
    """Service for generating XLSX exports from PDC and weekly plans"""

    @staticmethod
    def export_pdc(
        pdc: PDC,
        weekly_plans: Optional[List] = None,
        branding: Optional[BrandingConfig] = None,
    ) -> bytes:
        """
        Generate an XLSX document from PDC content with school branding.

        Args:
            pdc: PDC model instance
            weekly_plans: Optional list of weekly plan objects
            branding: Optional BrandingConfig with school branding

        Returns:
            XLSX file as bytes

        Raises:
            ImportError: If openpyxl is not available
        """
        if not OPENPYXL_AVAILABLE:
            raise ImportError(
                "openpyxl is not installed. "
                "Install it with: pip install openpyxl"
            )

        branding = branding or BrandingConfig()
        weekly_plans = weekly_plans or []

        # Create workbook with 3 sheets
        wb = Workbook()
        wb.remove(wb.active)  # Remove default sheet

        # Sheet 1: PDC Overview
        ws1 = wb.create_sheet("PDC Overview", 0)
        XlsxExportService._create_overview_sheet(ws1, pdc, branding)

        # Sheet 2: Weekly Plans
        ws2 = wb.create_sheet("Weekly Plans", 1)
        XlsxExportService._create_weekly_plans_sheet(ws2, pdc, weekly_plans, branding)

        # Sheet 3: Micro-Objectives
        ws3 = wb.create_sheet("Micro-Objectives", 2)
        XlsxExportService._create_objectives_sheet(ws3, pdc, branding)

        # Return bytes
        from io import BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()

    @staticmethod
    def _create_overview_sheet(ws, pdc: PDC, branding: BrandingConfig) -> None:
        """Create the PDC Overview sheet"""
        header_fill = PatternFill(
            start_color=hex_to_rgb(branding.header_color),
            end_color=hex_to_rgb(branding.header_color),
            fill_type="solid"
        )
        header_font = Font(bold=True, color="FFFFFF", size=12)
        border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin")
        )

        # Title
        ws["A1"] = branding.school_name
        ws["A1"].font = Font(bold=True, size=14, color=hex_to_rgb(branding.header_color))
        ws.merge_cells("A1:D1")

        ws["A2"] = "Plan de Desarrollo Curricular (PDC)"
        ws["A2"].font = Font(bold=True, size=11)
        ws.merge_cells("A2:D2")

        # Metadata
        row = 4
        ws[f"A{row}"] = "Subject:"
        ws[f"B{row}"] = pdc.subject or "N/A"
        row += 1

        ws[f"A{row}"] = "Grade Level:"
        ws[f"B{row}"] = pdc.grade_level or "N/A"
        row += 1

        ws[f"A{row}"] = "Trimester:"
        ws[f"B{row}"] = pdc.trimester
        row += 1

        ws[f"A{row}"] = "School Year:"
        ws[f"B{row}"] = pdc.school_year
        row += 1

        ws[f"A{row}"] = "Total Weekly Plans:"
        ws[f"B{row}"] = len(pdc.weekly_plans) if hasattr(pdc, "weekly_plans") else 0
        row += 1

        ws[f"A{row}"] = "Generated:"
        ws[f"B{row}"] = datetime.utcnow().strftime("%d/%m/%Y %H:%M:%S")
        row += 1

        # Adjust column widths
        ws.column_dimensions["A"].width = 20
        ws.column_dimensions["B"].width = 30

    @staticmethod
    def _create_weekly_plans_sheet(ws, pdc: PDC, weekly_plans: List, branding: BrandingConfig) -> None:
        """Create the Weekly Plans sheet with status and metrics"""
        header_fill = PatternFill(
            start_color=hex_to_rgb(branding.header_color),
            end_color=hex_to_rgb(branding.header_color),
            fill_type="solid"
        )
        header_font = Font(bold=True, color="FFFFFF")
        border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin")
        )

        # Headers
        headers = ["Week", "Status", "Momentos Total", "Word Count", "Last Updated"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Data rows (weeks 15-30)
        for week_num in range(15, 31):
            row = week_num - 14
            ws.cell(row=row + 1, column=1).value = f"Week {week_num}"

            # Find matching weekly plan
            plan = None
            if weekly_plans:
                plan = next((p for p in weekly_plans if getattr(p, "week", None) == week_num), None)

            if plan:
                status = getattr(plan, "status", "Draft")
                ws.cell(row=row + 1, column=2).value = status

                # Conditional formatting for status
                status_fill = PatternFill(
                    start_color="0066CC" if status == "Draft" else "00B050",
                    end_color="0066CC" if status == "Draft" else "00B050",
                    fill_type="solid"
                )
                ws.cell(row=row + 1, column=2).fill = status_fill
                ws.cell(row=row + 1, column=2).font = Font(color="FFFFFF")

                momentos_count = len(getattr(plan, "momentos", [])) if hasattr(plan, "momentos") else 0
                ws.cell(row=row + 1, column=3).value = momentos_count

                word_count = len(str(getattr(plan, "content", "")).split())
                ws.cell(row=row + 1, column=4).value = word_count

                # Conditional formatting for word count (red if >500)
                if word_count > 500:
                    ws.cell(row=row + 1, column=4).fill = PatternFill(
                        start_color="FF0000",
                        end_color="FF0000",
                        fill_type="solid"
                    )
                    ws.cell(row=row + 1, column=4).font = Font(color="FFFFFF")

                updated = getattr(plan, "updated_at", datetime.utcnow())
                ws.cell(row=row + 1, column=5).value = updated.strftime("%d/%m/%Y") if updated else "N/A"
            else:
                ws.cell(row=row + 1, column=2).value = "Not Created"
                ws.cell(row=row + 1, column=3).value = 0
                ws.cell(row=row + 1, column=4).value = 0

            # Apply borders
            for col in range(1, 6):
                ws.cell(row=row + 1, column=col).border = border
                ws.cell(row=row + 1, column=col).alignment = Alignment(horizontal="center")

        # Adjust column widths
        ws.column_dimensions["A"].width = 12
        ws.column_dimensions["B"].width = 15
        ws.column_dimensions["C"].width = 15
        ws.column_dimensions["D"].width = 12
        ws.column_dimensions["E"].width = 15

    @staticmethod
    def _create_objectives_sheet(ws, pdc: PDC, branding: BrandingConfig) -> None:
        """Create the Micro-Objectives sheet"""
        header_fill = PatternFill(
            start_color=hex_to_rgb(branding.header_color),
            end_color=hex_to_rgb(branding.header_color),
            fill_type="solid"
        )
        header_font = Font(bold=True, color="FFFFFF")
        border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin")
        )

        # Headers
        headers = ["Week", "Objective", "Description"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border

        # Add objectives from weekly plans
        row = 2
        if hasattr(pdc, "weekly_plans") and pdc.weekly_plans:
            for week_num, plan in enumerate(pdc.weekly_plans, 15):
                objectives = getattr(plan, "micro_objetivos", [])
                if objectives:
                    for obj in objectives[:3]:  # Limit to 3 per week
                        ws.cell(row=row, column=1).value = f"Week {week_num}"
                        ws.cell(row=row, column=2).value = getattr(obj, "name", "Objective")
                        ws.cell(row=row, column=3).value = getattr(obj, "description", "")
                        row += 1
                else:
                    ws.cell(row=row, column=1).value = f"Week {week_num}"
                    ws.cell(row=row, column=2).value = pdc.objetivo or "No objective defined"
                    row += 1

        # Adjust column widths
        ws.column_dimensions["A"].width = 12
        ws.column_dimensions["B"].width = 30
        ws.column_dimensions["C"].width = 40
