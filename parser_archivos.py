"""
parser_archivos.py — Parsers para los archivos fuente del Sistema Diario PRIA
==============================================================================
Parsea los 4 archivos que el admin sube y devuelve listas de dicts
listas para insertarse en la base de datos.

  parse_horarios(wb_bytes)      → list[dict]  para horario_docente
  parse_calendario(wb_bytes)    → list[dict]  para calendario_escolar
  parse_cronograma(wb_bytes)    → list[dict]  para actividades_cronograma
  parse_comisiones(docx_bytes)  → list[dict]  para comisiones_docente
"""

import io
import re
from datetime import datetime

import openpyxl
import docx as _docx


# ─────────────────────────────────────────────────────────────────────────────
# HELPERS COMPARTIDOS
# ─────────────────────────────────────────────────────────────────────────────


def _parse_time_range(s: str) -> tuple[str | None, str | None]:
    """Extrae (hora_inicio, hora_fin) de strings como '07:55 - 08:40'."""
    if not s:
        return None, None
    s = str(s).strip()

    def _pad(t: str) -> str:
        return f"0{t}" if len(t) == 4 else t

    m = re.search(r"(\d{1,2}:\d{2})\s*[-–]\s*(\d{1,2}:\d{2})", s)
    if m:
        return _pad(m.group(1)), _pad(m.group(2))
    m = re.search(r"(\d{1,2}:\d{2})", s)
    if m:
        return _pad(m.group(1)), None
    return None, None


def _classify_block(valor: str) -> tuple[str, str | None, str | None, str | None]:
    """
    Clasifica un valor de celda del horario.
    Devuelve (tipo_bloque, materia, nivel_grado, ubicacion).
    """
    v = valor.upper().strip()

    if not v or v == "NONE":
        return "vacio", None, None, None

    if "HORARIO DE INGRESO" in v:
        return "ingreso", None, None, None

    if "VIGILANCIA RECREO" in v or "GUARDIA RECREO" in v:
        if "PRIM" in v:
            return "vigilancia_recreo", None, None, "Área Primaria"
        if "SEC" in v:
            return "vigilancia_recreo", None, None, "Área Secundaria"
        return "vigilancia_recreo", None, None, "Patio"

    if "AT. PPFF" in v or "ATENCIÓN PPFF" in v or "ATENCION PPFF" in v:
        return "atencion_ppff", None, None, None

    if "PLANIFICACIÓN" in v or "PLANIFICACION" in v or ("API" in v and "PLANIF" in v):
        return "planificacion", None, None, None

    if "RECESO" in v:
        return "recreo_libre", None, None, None

    if "ACOMPAÑAMIENTO" in v or "ACOMPANAMIENTO" in v:
        # e.g. "ACOMPAÑAMIENTO INGLES" → clase de acompañamiento
        partes = v.replace("ACOMPAÑAMIENTO", "").replace("ACOMPANAMIENTO", "").strip()
        return "clase", f"Acompañamiento {partes.title()}", None, None

    # Bloque de clase — extraer materia y nivel/grado
    # Patrones: "TECNOLOGIA 2S", "HABILIDADES 5P", "LENGUAJE", "5S", "4S"
    # Nivel al final: dígito + letra (P=primaria, S=secundaria)
    nivel_match = re.search(r"\b(\d+[PS])\b", v)
    nivel = nivel_match.group(1) if nivel_match else None
    if nivel:
        materia_raw = v[: nivel_match.start()].strip().rstrip(" -")
    else:
        materia_raw = v

    # Si queda solo un código de grado (ej. "4S", "3S")
    if re.match(r"^\d+[PS]$", materia_raw.strip()):
        return "clase", materia_raw.strip().title(), nivel, None

    materia = (
        materia_raw.strip().title() if materia_raw.strip() else valor.strip().title()
    )
    return "clase", materia, nivel, None


# ─────────────────────────────────────────────────────────────────────────────
# PARSER 1 — HORARIOS (Excel, una hoja por docente)
# ─────────────────────────────────────────────────────────────────────────────

SKIP_SHEETS = {"HORARIO POR CURSO"}
DIAS_ES = ["lunes", "martes", "miercoles", "jueves", "viernes"]
DIAS_HEADER = ["LUNES", "MARTES", "MIÉRCOLES", "MIERCOLES", "JUEVES", "VIERNES"]


def parse_horarios(wb_bytes: bytes) -> list[dict]:
    """
    Parsea HORARIOS_CORREGIDOS.xlsx.
    Devuelve lista de dicts para horario_docente.

    Handles dual time columns (SEC + PRIM) in teacher sheets:
    - Headers row shows | SEC | PRIM | LUNES | MARTES | ...
    - Some rows belong to SEC level, others to PRIM level
    - Level is determined from the content cell (e.g. "3S" = SEC, "3P" = PRIM,
      or "VIGILANCIA RECREO PRIM HASTA 1" contains "PRIM")
    """
    wb = openpyxl.load_workbook(io.BytesIO(wb_bytes), data_only=True)
    records = []

    for sheet_name in wb.sheetnames:
        if sheet_name.strip().upper() in SKIP_SHEETS:
            continue
        ws = wb[sheet_name]
        nombre_hoja = sheet_name.strip().upper()

        # ── 1. Find header row (LUNES) and detect dual time columns ────────────
        header_row_idx = None
        lunes_col_idx = None
        has_dual_time = False  # True = this sheet has both SEC and PRIM time cols

        for r_idx, row in enumerate(ws.iter_rows(values_only=True)):
            for c_idx, cell in enumerate(row):
                if str(cell).upper().strip() in ("LUNES",):
                    header_row_idx = r_idx
                    lunes_col_idx = c_idx
                    break
            if header_row_idx is not None:
                break

        if header_row_idx is None or lunes_col_idx is None:
            continue

        # Detect dual time columns: check cell immediately left of LUNES
        # If it contains "PRIM" or "PRIMARIA" text, it's the PRIM time column header
        # → meaning SEC time column is one more to the left
        time_col_prim = None
        if lunes_col_idx > 0:
            left_header = str(
                ws.cell(header_row_idx + 1, lunes_col_idx).value or ""
            ).upper()
            # Actually check the header row itself (header_row_idx)
            hdr_left = str(
                ws.cell(header_row_idx, lunes_col_idx - 1).value or ""
            ).upper()
            if "PRIM" in hdr_left or "PRIMARIA" in hdr_left:
                has_dual_time = True
                time_col_prim = lunes_col_idx - 1
                # SEC time column is one further left
                time_col_sec = lunes_col_idx - 2
            else:
                time_col_sec = lunes_col_idx - 1

        # ── 2. Parse data rows ─────────────────────────────────────────────────
        for row_idx, row in enumerate(
            ws.iter_rows(min_row=header_row_idx + 2, values_only=True)
        ):
            # Choose correct time column based on level of this row's content
            tiempo_raw = None
            if has_dual_time and time_col_sec is not None:
                # Check the content cells (LUNES..VIERNES) to detect level
                # Level "PRIM" = primaria, any other indicator = secundaria
                row_level = None
                for d_offset in range(5):  # 5 day columns
                    col_idx = lunes_col_idx + d_offset
                    if col_idx < len(row) and row[col_idx]:
                        val = str(row[col_idx]).upper()
                        if "PRIM" in val and "RECREO" in val:
                            row_level = "PRIM"
                            break
                        elif "SEC" in val or val.endswith("S") or val.endswith("P"):
                            # Check for explicit nivel marker
                            import re as _re

                            m = _re.search(r"\b(\d+[PS])\b", val)
                            if m:
                                row_level = m.group(1)
                                break

                # Use PRIM time if row_level is PRIM (3P, 5P, PRIM keywords)
                if row_level and (
                    "P" in row_level.upper()
                    or "PRIM" in str(row.get(lunes_col_idx) or "").upper()
                ):
                    if time_col_prim is not None:
                        tiempo_raw = (
                            row[time_col_prim] if time_col_prim < len(row) else None
                        )
                else:
                    # Use SEC time (default to SEC for ambiguous rows)
                    tiempo_raw = row[time_col_sec] if time_col_sec < len(row) else None
            else:
                # Single time column
                time_col = lunes_col_idx - 1
                tiempo_raw = row[time_col] if time_col < len(row) else None

            if not tiempo_raw:
                continue
            tiempo_str = str(tiempo_raw).strip()
            if not any(c.isdigit() for c in tiempo_str):
                continue

            hora_inicio, hora_fin = _parse_time_range(tiempo_str)
            if not hora_inicio:
                continue

            for d_offset, dia in enumerate(DIAS_ES):
                col_idx = lunes_col_idx + d_offset
                if col_idx >= len(row):
                    continue
                valor = row[col_idx]
                if not valor:
                    continue
                valor_str = str(valor).strip()
                if not valor_str or valor_str.upper() == "NONE":
                    continue

                tipo, materia, nivel, ubicacion = _classify_block(valor_str)
                if tipo == "vacio":
                    continue

                records.append(
                    {
                        "nombre_hoja": nombre_hoja,
                        "dia_semana": dia,
                        "hora_inicio": hora_inicio,
                        "hora_fin": hora_fin,
                        "tipo_bloque": tipo,
                        "materia": materia,
                        "nivel_grado": nivel,
                        "ubicacion": ubicacion,
                        "valor_original": valor_str,
                    }
                )

    return records


# ─────────────────────────────────────────────────────────────────────────────
# PARSER 2 — CALENDARIO ESCOLAR (Excel, 3 hojas trimestrales)
# ─────────────────────────────────────────────────────────────────────────────


def _classify_evento(nombre: str) -> str:
    n = nombre.upper()
    if "FERIADO" in n:
        return "feriado"
    if "ACTO CÍVICO" in n or "ACTO CIVICO" in n or "CONMEMORATIV" in n:
        return "acto_civico"
    if "DESARROLLO CURRICULAR" in n or "AVANCE DE CONTENIDOS" in n:
        return "curricular"
    return "institucional"


def parse_calendario(wb_bytes: bytes) -> list[dict]:
    """
    Parsea 'Calendario Interno - LPS 2026_.xlsx'.
    Devuelve lista de dicts para calendario_escolar.
    """
    wb = openpyxl.load_workbook(io.BytesIO(wb_bytes), data_only=True)
    records = []
    seen = set()  # evitar duplicados por (fecha, nombre_evento)

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        # Detectar columnas: buscamos fila con 'DIA' o 'FECHA'
        # Estructura conocida: col 0 = SEMANAS, col 1 = DIA (datetime), col 2 = ITEM, col 3 = ACTIVIDADES, col 4 = OBS
        # Pero en T1 puede variar. Usamos búsqueda flexible: primera columna con datetime.

        for row in ws.iter_rows(values_only=True):
            # Buscar la celda datetime en la fila
            fecha_val = None
            fecha_col = None
            for c_idx, cell in enumerate(row):
                if isinstance(cell, datetime):
                    fecha_val = cell
                    fecha_col = c_idx
                    break

            if fecha_val is None:
                continue

            fecha_str = fecha_val.strftime("%Y-%m-%d")

            # Las columnas siguientes al DIA contienen ITEM, ACTIVIDADES, OBS
            item = (
                str(row[fecha_col + 1]).strip()
                if fecha_col + 1 < len(row) and row[fecha_col + 1]
                else ""
            )
            actividad = (
                str(row[fecha_col + 2]).strip()
                if fecha_col + 2 < len(row) and row[fecha_col + 2]
                else ""
            )
            obs = (
                str(row[fecha_col + 3]).strip()
                if fecha_col + 3 < len(row) and row[fecha_col + 3]
                else ""
            )

            # Limpiar 'None' strings
            item = "" if item in ("None", "FALSE", "False") else item
            actividad = "" if actividad in ("None", "FALSE", "False") else actividad
            obs = "" if obs in ("None", "FALSE", "False") else obs

            nombre_evento = item or actividad
            if not nombre_evento:
                continue

            key = (fecha_str, nombre_evento[:60])
            if key in seen:
                continue
            seen.add(key)

            records.append(
                {
                    "fecha": fecha_str,
                    "nombre_evento": nombre_evento,
                    "descripcion": actividad if item else "",
                    "responsable": obs,
                    "tipo": _classify_evento(nombre_evento),
                }
            )

    return records


# ─────────────────────────────────────────────────────────────────────────────
# PARSER 3 — CRONOGRAMA SEMANAL DOCENTE (Excel, una hoja por semana)
# ─────────────────────────────────────────────────────────────────────────────


def parse_cronograma(wb_bytes: bytes) -> list[dict]:
    """
    Parsea 'CRONOGRAMA DE ACTIVIDADES SEMANAS DOCENTE.xlsx'.
    Devuelve lista de dicts para actividades_cronograma.
    """
    wb = openpyxl.load_workbook(io.BytesIO(wb_bytes), data_only=True)
    records = []
    seen = set()

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        current_date = None

        for row in ws.iter_rows(values_only=True):
            if not any(c for c in row):
                continue

            # Buscar datetime en la fila para actualizar fecha actual
            for cell in row:
                if isinstance(cell, datetime):
                    current_date = cell.strftime("%Y-%m-%d")
                    break

            if not current_date:
                continue

            # Columnas: DIA(0-1), HORA(1-2), ACTIVIDAD(2-3), A CARGO(3-4)
            # Formato varía — buscar la celda de hora y actividad
            hora_str = None
            actividad = None
            a_cargo = None

            # Intentar diferentes offsets según estructura detectada
            # Estructura conocida col 1 = DIA/datetime, col 2 = HORA, col 3 = ACTIVIDAD, col 4 = A CARGO
            if len(row) >= 4:
                hora_candidate = row[2] if len(row) > 2 else None
                actividad_candidate = row[3] if len(row) > 3 else None
                cargo_candidate = row[4] if len(row) > 4 else None

                hora_str = str(hora_candidate).strip() if hora_candidate else None
                actividad = (
                    str(actividad_candidate).strip() if actividad_candidate else None
                )
                a_cargo = str(cargo_candidate).strip() if cargo_candidate else None

            if not actividad or actividad in ("None", "ACTIVIDAD", ""):
                continue
            if "CRONOGRAMA" in actividad.upper() or "DÍA" in actividad.upper():
                continue

            hora_inicio, hora_fin = _parse_time_range(hora_str or "")
            a_cargo = "" if a_cargo in ("None", None) else a_cargo

            key = (current_date, actividad[:60], hora_inicio or "")
            if key in seen:
                continue
            seen.add(key)

            records.append(
                {
                    "fecha": current_date,
                    "hora_inicio": hora_inicio,
                    "hora_fin": hora_fin,
                    "actividad": actividad,
                    "a_cargo_de": a_cargo,
                }
            )

    return records


# ─────────────────────────────────────────────────────────────────────────────
# PARSER 4 — COMISIONES DOCENTES (Word .docx, tabla)
# ─────────────────────────────────────────────────────────────────────────────


def parse_comisiones(docx_bytes: bytes) -> list[dict]:
    """
    Parsea 'COMISIONES DOCENTES LPS 2026.docx'.
    Devuelve lista de dicts para comisiones_docente.
    """
    doc = _docx.Document(io.BytesIO(docx_bytes))
    records = []

    for table in doc.tables:
        for row in table.rows:
            cells = [c.text.strip() for c in row.cells]
            if len(cells) < 2:
                continue

            comision_nombre = cells[0].strip()
            miembros_raw = cells[1].strip()
            funciones = cells[2].strip() if len(cells) > 2 else ""

            # Saltar fila de encabezado (celdas fusionadas con mismo texto)
            if comision_nombre == miembros_raw:
                continue
            if not comision_nombre or not miembros_raw:
                continue
            # Saltar filas que no son comisiones reales
            if "LPS" in comision_nombre.upper() or "DOCENTE" in comision_nombre.upper():
                continue

            # Parsear lista de miembros
            # Formato: "1.       Adela Vargas\n2.       Glen Vargas - Festival\n..."
            for line in miembros_raw.split("\n"):
                line = line.strip()
                if not line:
                    continue
                # Eliminar prefijo numérico
                m = re.match(r"[\d]+[.)]\s*(.+)", line)
                nombre_full = m.group(1).strip() if m else line
                # Tomar solo el nombre (antes de guión si hay rol)
                nombre = re.split(r"\s+[-–]\s+", nombre_full)[0].strip()
                if not nombre or len(nombre) < 3:
                    continue

                # nombre_hoja = primer apellido o primera palabra en mayúsculas
                # Formato "Nombre Apellido" → usamos el primer nombre como hoja
                partes = nombre.split()
                nombre_hoja = partes[0].upper() if partes else nombre.upper()

                records.append(
                    {
                        "nombre_docente": nombre,
                        "nombre_hoja": nombre_hoja,
                        "comision": comision_nombre,
                        "funciones": funciones,
                    }
                )

    return records


# ─────────────────────────────────────────────────────────────────────────────
# PARSER 5 — VIGILANCIAS RECREO (PDF, page 2 grid)
# ─────────────────────────────────────────────────────────────────────────────


def _normalize_teacher_name(name: str) -> str:
    """
    Normalize a teacher name from the PDF to match DB nombre_hoja.
    PDF uses first names (Angélica, Yamile), DB uses uppercase first names (ANGÉLICA, YAMILE).
    Also handles full names that the PDF sometimes includes.
    """
    import unicodedata

    s = str(name).strip()
    # Remove accents
    s = "".join(
        c for c in unicodedata.normalize("NFD", s) if unicodedata.category(c) != "Mn"
    )
    # Uppercase and collapse spaces
    s = " ".join(s.upper().split())
    return s


def _split_teachers(cell_text: str) -> list[str]:
    """
    Split a cell containing multiple teachers into individual names.
    Handles: "Angélica - Yamile - Melani", "SEBASTIAN + ANTONIO", "Glen (10:25)"
    Returns list of cleaned first-name tokens.
    """
    if not cell_text or cell_text.strip() in ("", "OCUPADO"):
        return []
    text = cell_text.strip()
    # Split on " - ", " + ", or " / "
    parts = re.split(r"\s*[-+]\s*|\s*/\s*", text)
    teachers = []
    for p in parts:
        p = p.strip()
        if not p or p.upper() == "OCUPADO":
            continue
        # Remove time overrides like "(10:25)"
        p = re.sub(r"\(\d{2}:\d{2}\)", "", p).strip()
        if not p:
            continue
        teachers.append(p)
    return teachers


def parse_vigilancia_pdf(pdf_bytes: bytes) -> list[dict]:
    """
    Parse 'ROL DE VIGILANCIA _ RECREOS 2026.pdf' — page 2 recreational vigilances.

    Grid structure:
        Columns: LUNES, MARTES, MIERCOLES, JUEVES, VIERNES
        Rows: PRIMARIA R1 (10:10-10:30), PRIMARIA R2 (12:00-12:15),
              SECUNDARIA R1 (09:25-09:40), SECUNDARIA R2 (11:10-11:25)
        Cell content: "Angélica - Yamile" or "Sebastian (10:25)" or "OCUPADO"

    Returns list of dicts:
        nombre_hoja, nivel, dia_semana, recreo_num, zona,
        hora_inicio, hora_fin, nota (source)
    """
    try:
        import pdfplumber
    except ImportError:
        raise RuntimeError(
            "pdfplumber is required for PDF parsing. Install with: pip install pdfplumber"
        )

    records = []
    zonas_primaria = [
        "PARQUE",
        "ESCALERAS B E INGRESO AL COLISEO",
        "COLISEO",
        "KIOSCO",
        "PATIO CENTRAL",
        "BIBLIOTECA",
    ]
    zonas_secundaria = zonas_primaria + ["ESCALERAS A"]

    RECESO_SLOTS = {
        "primaria_r1": ("10:10", "10:30"),
        "primaria_r2": ("12:00", "12:15"),
        "secundaria_r1": ("09:25", "09:40"),
        "secundaria_r2": ("11:10", "11:25"),
    }

    with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
        if len(pdf.pages) < 2:
            raise ValueError(f"PDF has {len(pdf.pages)} pages, expected at least 2")

        page = pdf.pages[1]  # page 2 (0-indexed)
        tables = page.extract_tables()

    if not tables:
        raise ValueError("No tables found in page 2 of the PDF")

    # Identify which table is PRIMARIA and which is SECUNDARIA
    primaria_table = None
    sec_table = None

    for tbl in tables:
        if not tbl or len(tbl) < 3:
            continue
        header = [str(c or "").strip().upper() for c in tbl[0]]
        first_col = header[0] if header else ""
        if any(z in first_col for z in zonas_primaria):
            primaria_table = tbl
            continue
        if (
            first_col.startswith("LUNES")
            or first_col.startswith("DIA")
            or any(z in " ".join(header) for z in zonas_secundaria)
        ):
            if "PARQUE" in " ".join(header) or "SECUNDARIA" in " ".join(header):
                sec_table = tbl

    # Fallback: first table = primaria, second = sec
    if primaria_table is None and len(tables) >= 1:
        primaria_table = tables[0]
    if sec_table is None and len(tables) >= 2:
        sec_table = tables[1]

    def _parse_grid(table, nivel, slots):
        """Parse a vigilancia grid table into records."""
        if table is None:
            return
        if len(table) < 2:
            return

        headers = [str(c or "").strip().upper() for c in table[0]]
        # Map column index → day name
        dias_map = {}
        for c_idx, h in enumerate(headers):
            h_clean = h.replace("PRIMARIA", "").replace("SECUNDARIA", "").strip()
            for d in ["LUNES", "MARTES", "MIÉRCOLES", "MIERCOLES", "JUEVES", "VIERNES"]:
                if d in h_clean:
                    dias_map[c_idx] = d.lower()
                    break

        # Rows: [slot_label, zone1_teacher, zone2_teacher, ...]
        for row in table[1:]:
            if not row or len(row) < 2:
                continue
            slot_label = str(row[0] or "").strip()
            if not slot_label:
                continue

            slot_upper = slot_label.upper()
            slot_key = None
            if "R1" in slot_upper and "PRIM" in slot_upper:
                slot_key = "primaria_r1"
            elif "R2" in slot_upper and "PRIM" in slot_upper:
                slot_key = "primaria_r2"
            elif "R1" in slot_upper and "SEC" in slot_upper:
                slot_key = "secundaria_r1"
            elif "R2" in slot_upper and "SEC" in slot_upper:
                slot_key = "secundaria_r2"

            h_ini, h_fin = None, None
            if slot_key and slot_key in RECESO_SLOTS:
                h_ini, h_fin = RECESO_SLOTS[slot_key]

            # Time override like (10:25)
            override_match = re.search(r"\((\d{2}:\d{2})\)", slot_label)
            if override_match:
                h_ini = override_match.group(1)

            if h_ini is None:
                continue

            # Zone names from header columns
            zone_names = headers[1:] if len(headers) > 1 else []
            recreo_num = "1" if "R1" in slot_upper else "2"

            for c_idx in range(1, len(row)):
                if c_idx >= len(row):
                    break
                cell_text = str(row[c_idx] or "").strip()
                if not cell_text or cell_text.upper() == "OCUPADO":
                    continue

                zona = zone_names[c_idx - 1] if c_idx - 1 < len(zone_names) else ""
                dia = dias_map.get(c_idx, None)
                if dia is None:
                    continue

                teachers = _split_teachers(cell_text)
                if not teachers:
                    continue

                for teacher in teachers:
                    records.append(
                        {
                            "nombre_hoja": _normalize_teacher_name(teacher),
                            "nivel": nivel,
                            "dia_semana": dia,
                            "recreo_num": recreo_num,
                            "zona": zona,
                            "hora_inicio": h_ini,
                            "hora_fin": h_fin,
                            "nota": f"pdf:{slot_label.strip()}",
                        }
                    )

    _parse_grid(primaria_table, "primaria", RECESO_SLOTS)
    _parse_grid(sec_table, "secundaria", RECESO_SLOTS)

    return records
